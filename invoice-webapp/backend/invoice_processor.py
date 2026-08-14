import os
import xml.etree.ElementTree as ET
import re
import pdfplumber
import copy
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Tiện ích chuyển đổi số thành chữ tiếng Việt
def number_to_words_vn(n):
    if n == 0: return "Không đồng"
    units = ["", "nghìn", "triệu", "tỷ", "nghìn tỷ", "triệu tỷ"]
    digits = ["không", "một", "hai", "ba", "bốn", "năm", "sáu", "bảy", "tám", "chín"]
    
    def read_group(group):
        h, t, u = group // 100, (group % 100) // 10, group % 10
        res = []
        if h > 0 or t > 0 or u > 0:
            res.append(digits[h] + " trăm")
            if t == 0 and u > 0: res.append("linh")
            elif t == 1: res.append("mười")
            elif t > 1: res.append(digits[t] + " mươi")
            if u == 1 and t > 1: res.append("mốt")
            elif u == 5 and t > 0: res.append("lăm")
            elif u > 0: res.append(digits[u])
        return " ".join(res).strip()
    
    s = str(n)
    groups = []
    while len(s) > 0:
        groups.append(int(s[-3:]))
        s = s[:-3]
    
    words = []
    for i, group in enumerate(groups):
        if group > 0:
            group_word = read_group(group)
            if i == len(groups) - 1 and group < 100 and len(groups) > 1:
                group_word = group_word.replace(" không trăm", "")
                if group_word.startswith("linh "): group_word = group_word[5:]
            words.append(group_word + " " + units[i])
            
    res = " ".join(reversed(words)).strip()
    res = res.capitalize() + " đồng chẵn./."
    res = res.replace("  ", " ")
    if res.startswith(" không trăm "): res = res[12:].capitalize()
    return res

def parse_xml_invoice(filepath):
    try:
        tree = ET.parse(filepath)
        root = tree.getroot()
        
        # Strip namespace for easier searching
        for elem in root.iter():
            if '}' in elem.tag:
                elem.tag = elem.tag.split('}', 1)[1]
        
        # === Basic fields ===
        invoice_no = root.find('.//SHDon')
        invoice_no = invoice_no.text if invoice_no is not None else ""
        
        date = root.find('.//NLap')
        if date is not None and date.text:
            date_str = date.text[:10]
            if re.match(r'\d{4}-\d{2}-\d{2}', date_str):
                parts = date_str.split('-')
                date = f"{parts[2]}/{parts[1]}/{parts[0]}"
            else:
                date = date_str
        else:
            date = ""
        
        # Provider: try NBan/Ten, check .text is not None
        provider_elem = root.find('.//NBan/Ten')
        if provider_elem is None or provider_elem.text is None:
            provider_elem = root.find('.//Ten')
        provider = provider_elem.text if (provider_elem is not None and provider_elem.text) else "Nhà cung cấp"
        
        # Cố gắng tìm giá trị sau thuế (Tổng tiền thanh toán)
        amount_tags = ['TgTTTBSo', 'TongTienThanhToan', 'TongTien', 'TienThanhtoan']
        amount_val = 0
        for tag in amount_tags:
            elem = root.find(f'.//{tag}')
            if elem is not None and elem.text:
                try:
                    amount_val = float(elem.text)
                    if amount_val > 0:
                        break
                except ValueError:
                    pass
        amount = amount_val
        
        # === Collect THHDVu and TTin ===
        thhdvu_texts = []
        for t in root.findall('.//THHDVu'):
            if t.text and t.text.strip():
                thhdvu_texts.append(t.text.strip())
        
        ttin_data = {}
        for ttin in root.findall('.//TTin'):
            ttruong = ttin.find('TTruong')
            dlieu = ttin.find('DLieu')
            if ttruong is not None and ttruong.text and dlieu is not None and dlieu.text:
                ttin_data[ttruong.text] = dlieu.text
        
        # === Determine type ===
        prov_upper = provider.upper()
        is_airline = "VIETJET" in prov_upper or "VIETNAM AIRLINES" in prov_upper
        is_taxi = any(k in prov_upper for k in ("GSM", "XANH", "DI CHUYỂN XANH", "GREEN CAR", "PHÚ HOÀNG", "SACO", "ĐẠI THÀNH", "TỴ MÙI"))
        
        # --- AIRLINE ---
        if is_airline:
            pnr = ""
            routes = []
            for txt in thhdvu_texts:
                if re.match(r'^[A-Z0-9]{5,7}$', txt):
                    pnr = txt
                    break
            for val in ttin_data.values():
                if re.match(r'^[A-Z]{3}-[A-Z]{3}$', val) and val not in routes:
                    routes.append(val)
            filename = os.path.basename(filepath)
            m = re.search(r'([A-Z]{3}-[A-Z]{3})', filename)
            if m and m.group(1) not in routes:
                routes.append(m.group(1))
            
            airline = "Vietjet Air" if "VIETJET" in prov_upper else "Vietnam Airlines"
            if pnr and routes:
                provider = f"{airline} [PNR: {pnr} | {', '.join(routes)}]"
            elif pnr:
                provider = f"{airline} [PNR: {pnr}]"
            elif routes:
                provider = f"{airline} [{', '.join(routes)}]"
            else:
                provider = airline
        
        # --- TAXI ---
        elif is_taxi:
            pickup = ttin_data.get('PICK_UP_ADDRESS', '') or ttin_data.get('SENDER_NAME', '')
            dropoff = ttin_data.get('DROP_OFF_ADDRESS', '') or ttin_data.get('SHIPPING_ADDRESS', '')
            if not pickup:
                pickup = ttin_data.get('CustomField2', '')
            if not dropoff:
                dropoff = ttin_data.get('CustomField3', '')
            
            if not pickup and not dropoff:
                for txt in thhdvu_texts:
                    m = re.search(r'Điểm đón:\s*(.*?)\s*[-–]\s*Điểm trả:\s*(.*)', txt)
                    if m:
                        pickup, dropoff = m.group(1).rstrip(' .'), m.group(2).rstrip(' .')
                        break
                    m = re.search(r'Điểm đón:\s*(.*?)\s*[.]\s*Điểm đến:\s*(.*)', txt)
                    if m:
                        pickup, dropoff = m.group(1).rstrip(' .'), m.group(2).rstrip(' .')
                        break
                    m = re.search(r'(?:taxi|xe|Cước)\s+.*?(?:BSX.*?)?[-–]?\s*Điểm đón:\s*(.*?)[-–]\s*Điểm.*?:\s*(.*)', txt, re.IGNORECASE)
                    if m:
                        pickup, dropoff = m.group(1).rstrip(' .'), m.group(2).rstrip(' .')
                        break
                    m = re.search(r'từ\s+(.*?)\s+đến\s+(.*)', txt, re.IGNORECASE)
                    if m:
                        pickup, dropoff = m.group(1).rstrip(' .,'), m.group(2).rstrip(' .,')
                        break
            
            def clean_addr(addr):
                if not addr:
                    return ""
                return addr.split(',')[0].strip()
            
            taxi_map = {
                "GSM": "Xanh SM", "XANH": "Xanh SM", "DI CHUYỂN XANH": "Xanh SM",
                "GREEN CAR": "Green Car", "PHÚ HOÀNG": "Phú Hoàng",
                "SACO": "Saco", "ĐẠI THÀNH": "Đại Thành Công", "TỴ MÙI": "Tỵ Mùi"
            }
            short_name = provider
            for key, name in taxi_map.items():
                if key in prov_upper:
                    short_name = name
                    break
            
            if pickup or dropoff:
                route_str = f" ({clean_addr(pickup)} -> {clean_addr(dropoff)})"
                if len(route_str) > 80:
                    route_str = route_str[:77] + "...)"
                provider = short_name + route_str
            else:
                provider = short_name
        
        # --- OTHER (restaurant, hospital, hotel, etc.) ---
        else:
            descs = []
            for txt in thhdvu_texts:
                if re.match(r'^[A-Z0-9]{5,}$', txt):
                    continue
                if any(skip in txt for skip in ('Cước phí vận chuyển mã', 'Chiết khấu thương mại', 'Phí nền tảng mã', 'Các loại phí dịch vụ')):
                    continue
                descs.append(txt)
            
            if descs:
                desc = descs[0]
                if len(desc) > 60:
                    desc = desc[:57] + "..."
                provider = f"{provider} ({desc})"
            
            filename = os.path.basename(filepath)
            m = re.search(r'([A-Z0-9]{5,6})_([A-Z]{3}-[A-Z]{3})', filename)
            if m:
                pnr, route = m.groups()
                route_info = f" [PNR: {pnr} | {route}]"
                if route_info not in provider:
                    provider += route_info
        
        return {
            'date': date[:10] if date else "",
            'provider': provider,
            'invoice_no': invoice_no,
            'amount': amount,
            'source': filepath
        }
    except Exception as e:
        print(f"Error parsing XML {filepath}: {e}")
        return None

def extract_pdf_data(filepath):
    # This is a generic fallback/heuristic PDF parser
    try:
        with pdfplumber.open(filepath) as pdf:
            text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
                
        amount = 0
        amounts = []
        # Tìm tất cả các số tiền liên quan đến tổng cộng/thành tiền
        for match in re.finditer(r'(?i)(tổng cộng tiền thanh toán|tổng tiền thanh toán|tổng cộng|total payment|total amount|tổng tiền|thành tiền).*?([0-9]{1,3}(?:[.,][0-9]{3})+|[0-9]{4,})', text):
            amount_str = match.group(2).replace(',', '').replace('.', '')
            if amount_str.isdigit():
                amounts.append(float(amount_str))
        
        # Giá trị sau thuế luôn là giá trị lớn nhất (bao gồm cả thuế)
        if amounts:
            amount = max(amounts)
                
        date = ""
        date_match = re.search(r'(\d{2}[/-]\d{2}[/-]\d{4})', text)
        if date_match:
            date = date_match.group(1).replace('-', '/')
        
        # Fallback: Vietnamese format "DD Tháng MM Năm YYYY"
        if not date:
            vn_date = re.search(r'(\d{1,2})\s*Tháng.*?(\d{1,2})\s*Năm.*?(\d{4})', text)
            if vn_date:
                d, m, y = vn_date.groups()
                date = f"{int(d):02d}/{int(m):02d}/{y}"
        
        # Fallback: YYYYMMDD from filename prefix
        if not date:
            filename = os.path.basename(filepath)
            fname_date = re.match(r'(\d{4})(\d{2})(\d{2})', filename)
            if fname_date:
                y, m, d = fname_date.groups()
                date = f"{d}/{m}/{y}"
            
        # Extract Route/PNR from filename or text if available
        filename = os.path.basename(filepath)
        route_info = ""
        m = re.search(r'([A-Z0-9]{5,6})_([A-Z]{3}-[A-Z]{3})', filename)
        if m:
            pnr, route = m.groups()
            route_info = f" [PNR: {pnr} | {route}]"
        else:
            m2 = re.search(r'([A-Z]{3}-[A-Z]{3})', filename)
            if m2: route_info = f" [{m2.group(1)}]"
        
        # Extract invoice number from PDF text
        pdf_invoice_no = "N/A"
        inv_match = re.search(r'Số[:\s]*(\d+)', text)
        if inv_match:
            pdf_invoice_no = inv_match.group(1)
        
        # Detect provider from PDF content
        provider = "Hóa đơn/Biên lai (PDF)"
        text_lower = text.lower()
        
        if "vietjet" in text_lower:
            provider = "Vietjet Air" + route_info
        elif "vietnam airlines" in text_lower:
            provider = "Vietnam Airlines" + route_info
        elif "pvi" in text_lower:
            provider = "Bảo hiểm PVI"
        elif "xác nhận đặt phòng" in text_lower or "booking" in filename.lower():
            # Booking confirmation — not a VAT invoice, extract hotel name
            hotel_match = re.search(r'([\w\s]+Hotel|[\w\s]+Resort)', text)
            if hotel_match:
                provider = f"Đặt phòng KS {hotel_match.group(1).strip()}"
            else:
                provider = "Đặt phòng khách sạn"
        else:
            # Try to extract seller name from "Đơn vị bán hàng: XXX"
            seller = ""
            seller_match = re.search(r'(?i)(?:Đơn vị bán hàng|Tên người bán|Tên đơn vị bán|Người bán)[^\n:]*:\s*(.*?)(?:\n|Mã số thuế|Địa chỉ)', text)
            if seller_match:
                seller = seller_match.group(1).strip()
                if seller:
                    provider = seller

            # Always try to get service description from the item lines
            item_match = re.search(r'(?i)(?:Tên hàng hóa|Diễn giải|Tên dịch vụ|Nội dung|Hàng hoá, dịch vụ).*?\n((?:.*?\n){1,5})', text)
            desc = ""
            if item_match:
                lines = item_match.group(1).split('\n')
                for d in lines:
                    d = d.strip()
                    if not d or len(d) < 4: continue
                    # Bỏ qua các dòng tiêu đề bảng hoặc tiếng Anh
                    if re.match(r'^(?:stt|số thứ tự|tên|đơn vị tính|số lượng|đơn giá|mã|\(|\[)', d.lower()): continue
                    if "name of goods" in d.lower() or "description" in d.lower() or "hàng(seller)" in d.lower(): continue
                    # Bỏ qua các dòng chỉ chứa số đánh dấu cột (ví dụ: 1 2 3 4 5 6 = 4 x 5)
                    if re.match(r'^[\d\s=xX\.\,\+\-]+$', d): continue
                    
                    # Bỏ số thứ tự đầu dòng (ví dụ "1 Dịch vụ lưu trú" -> "Dịch vụ lưu trú")
                    clean_d = re.sub(r'^\d+[\.\s]+', '', d)
                    desc = f" ({clean_d[:60]}...)" if len(clean_d) > 60 else f" ({clean_d})"
                    break
            
            if desc:
                provider = provider + desc
        return {
            'date': date,
            'provider': provider,
            'invoice_no': pdf_invoice_no,
            'amount': amount,
            'source': filepath
        }
    except Exception as e:
        print(f"Error parsing PDF {filepath}: {e}")
        return None

def process_invoices(input_dir: str, template_path: str, output_path: str):

    # === PHASE 0: Extract ZIP files containing XML/PDF ===
    import zipfile
    for root_dir, _, files in os.walk(input_dir):
        for file in files:
            if not file.lower().endswith('.zip'):
                continue
            zip_path = os.path.join(root_dir, file)
            try:
                with zipfile.ZipFile(zip_path, 'r') as zf:
                    members = zf.infolist()
                    if len(members) > 100:
                        raise ValueError("ZIP exceeds 100 entries")
                    if sum(member.file_size for member in members) > 100 * 1024 * 1024:
                        raise ValueError("ZIP exceeds 100 MiB after extraction")
                    for member_info in members:
                        member = member_info.filename
                        member_lower = member.lower()
                        if member_lower.endswith('.xml') or member_lower.endswith('.pdf'):
                            safe_name = os.path.basename(member)
                            if not safe_name or member_info.file_size > 20 * 1024 * 1024:
                                continue
                            target = os.path.join(root_dir, safe_name)
                            if not os.path.exists(target):
                                with zf.open(member_info) as src, open(target, 'wb') as dst:
                                    shutil.copyfileobj(src, dst, length=1024 * 1024)
            except Exception as e:
                print(f"Warning: Could not extract {zip_path}: {e}")

    invoices = []
    
    # Helper: extract PNR from filename (handles both mid and end positions)
    def extract_pnr(filename):
        # Pattern 1: PNR between underscores: _TJE2CX_
        m = re.search(r'[_-]([A-Z0-9]{5,6})[_]', filename)
        if m: return m.group(1)
        # Pattern 2: PNR at end before extension: _4XBPDR.pdf
        m = re.search(r'[_-]([A-Z0-9]{5,8})\.\w+$', filename)
        if m and not m.group(1).isdigit(): return m.group(1)
        # Pattern 3: PNR after dash: Itinerary-GARN8B.pdf
        m = re.search(r'[Ii]tinerary-([A-Z0-9]{5,8})\.\w+$', filename)
        if m: return m.group(1)
        return None
    
    # Collect all Receipt/XML files' PNRs for cross-reference
    receipt_pnrs = set()
    for root_dir, _, files in os.walk(input_dir):
        for file in files:
            fname_lower = file.lower()
            if 'receipt' in fname_lower and (fname_lower.endswith('.xml') or fname_lower.endswith('.pdf')):
                pnr = extract_pnr(file)
                if pnr: receipt_pnrs.add(pnr)
            elif fname_lower.endswith('.xml') and 'itinerary' not in fname_lower:
                pnr = extract_pnr(file)
                if pnr: receipt_pnrs.add(pnr)
    
    print(f"Scanning directory: {input_dir}")
    for root_dir, _, files in os.walk(input_dir):
        for file in files:
            filepath = os.path.join(root_dir, file)
            fname_lower = file.lower()
            
            if fname_lower.endswith('.xml'):
                data = parse_xml_invoice(filepath)
                if data and data['amount'] > 0:
                    invoices.append(data)
            elif fname_lower.endswith('.pdf'):
                # Skip PDF if corresponding XML exists (same base name)
                base_no_ext = filepath[:-4]
                if os.path.exists(base_no_ext + '.xml') or os.path.exists(base_no_ext + '.XML'):
                    continue
                
                # Skip Itinerary/copy PDFs — they are NOT invoices
                if 'itinerary' in fname_lower or ' copy' in fname_lower:
                    continue
                
                # Skip Insurance PDFs if a Receipt with same PNR exists (avoid counting twice)
                if fname_lower.startswith('ins') or '_ins-' in fname_lower or '_ins_' in fname_lower:
                    pnr = extract_pnr(file)
                    if pnr and pnr in receipt_pnrs:
                        continue
                
                data = extract_pdf_data(filepath)
                if data and data['amount'] > 0:
                    invoices.append(data)
    
    # === Post-collection deduplication ===
    # Strategy: prefer XML invoices over PDF ones when both exist.
    # Step 1: Build a set of (amount, date) pairs from XML-sourced invoices
    xml_amount_date = set()
    for inv in invoices:
        if inv['source'].lower().endswith('.xml'):
            xml_amount_date.add((inv['amount'], inv['date']))
    
    # Step 2: Remove PDF duplicates where XML exists with same amount+date
    filtered = []
    for inv in invoices:
        if inv['source'].lower().endswith('.pdf'):
            if (inv['amount'], inv['date']) in xml_amount_date:
                continue  # XML version exists, skip PDF
        filtered.append(inv)
    
    # Step 3: Remove remaining exact duplicates by invoice_no+amount
    seen = set()
    unique_invoices = []
    for inv in filtered:
        # Use invoice_no if it looks real (>3 digits), else use provider+amount+date
        inv_no = inv.get('invoice_no', '')
        if inv_no and inv_no != 'N/A' and len(inv_no) > 3:
            key = f"{inv_no}|{inv['amount']}"
        else:
            key = f"{inv['provider']}|{inv['amount']}|{inv['date']}"
        
        if key not in seen:
            seen.add(key)
            unique_invoices.append(inv)
    
    removed = len(invoices) - len(unique_invoices)
    if removed > 0:
        print(f"Removed {removed} duplicate(s).")
    invoices = unique_invoices
                    
    def date_sort_key(inv):
        d = inv['date']
        if not d: return ""
        # Convert DD/MM/YYYY -> YYYY/MM/DD for proper sorting
        parts = d.split('/')
        if len(parts) == 3:
            return f"{parts[2]}/{parts[1]}/{parts[0]}"
        return d
    invoices.sort(key=date_sort_key)
    
    print(f"Found {len(invoices)} valid invoices.")
    if len(invoices) == 0:
        raise ValueError("Không tìm thấy hoá đơn hợp lệ nào trong các file đã tải lên.")
        
    total_sum = sum(inv['amount'] for inv in invoices)
    print(f"Total Amount: {total_sum:,.0f} VND")
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Không tìm thấy file mẫu (template) tại {template_path}")
        
    wb = load_workbook(template_path)
    ws = wb.active
    
    # === AUTO-FILL: Row 5 — Ngày lập ĐNTT (today) ===
    from datetime import datetime
    today = datetime.now()
    ws.cell(row=5, column=5, value=f"Ngày {today.day:02d} tháng {today.month:02d} năm {today.year}")
    
    # === AUTO-FILL: Row 11 — Nội dung thanh toán (tổng quan + khoảng thời gian) ===
    min_date = invoices[0]['date']   # already sorted
    max_date = invoices[-1]['date']
    
    # Detect expense categories from invoice data
    categories = []
    has_flight = any(k in inv['provider'].upper() for inv in invoices for k in ('VIETJET', 'VIETNAM AIRLINES'))
    has_taxi = any(k in inv['provider'].upper() for inv in invoices for k in ('XANH SM', 'TAXI', 'GREEN CAR', 'PHÚ HOÀNG', 'TỴ MÙI', 'SACO'))
    has_hotel = any(k in inv['provider'].upper() for inv in invoices for k in ('KHÁCH SẠN', 'HOTEL', 'RESORT', 'THANH TÂN'))
    
    if has_taxi: categories.append('đi lại')
    if has_flight: categories.append('vé máy bay')
    if has_hotel: categories.append('khách sạn')
    # Add catch-all if there are other expenses
    other_count = len(invoices) - sum(1 for inv in invoices if any(
        k in inv['provider'].upper() for k in ('VIETJET', 'VIETNAM AIRLINES', 'XANH SM', 'GREEN CAR', 
        'PHÚ HOÀNG', 'TỴ MÙI', 'SACO', 'KHÁCH SẠN', 'HOTEL', 'RESORT', 'THANH TÂN')))
    if other_count > 0 and not categories:
        categories.append('chi phí')
    
    # Extract destination from folder name (e.g. "20260609 - VIETNAM-LAO" → "VIETNAM-LAO")
    folder_name = os.path.basename(input_dir)
    destination = ""
    if ' - ' in folder_name:
        destination = folder_name.split(' - ', 1)[1]
    
    content = f"Chi phí {', '.join(categories)}" if categories else "Chi phí"
    if destination:
        content += f" công tác {destination}"
    content += f" từ {min_date} đến {max_date}"
    ws.cell(row=11, column=4, value=content)
    
    start_row = 14
    
    # === PHASE 1: Unmerge ALL cells from start_row downward ===
    # This prevents merge corruption when inserting/deleting rows
    merges_to_remove = []
    for merge in ws.merged_cells.ranges:
        if merge.min_row >= start_row:
            merges_to_remove.append(merge)
    for merge in merges_to_remove:
        ws.unmerge_cells(str(merge))
    
    # === PHASE 2: Find and remove old data rows ===
    # Data rows: from row 14 until "Tổng cộng" row
    summary_row = None
    for row in range(start_row, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1).value
        c6 = ws.cell(row=row, column=6).value
        if c1 and isinstance(c1, str) and "tổng" in c1.lower():
            summary_row = row
            break
        if c6 and isinstance(c6, str) and "SUM" in c6.upper():
            summary_row = row
            break
    
    old_data_end = (summary_row - 1) if summary_row else start_row
    old_rows_count = max(0, old_data_end - start_row + 1)
    
    if old_rows_count > 0:
        # Clear all cell values in old data range
        for row in range(start_row, old_data_end + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None
        
        # Delete extra rows (keep exactly 1 row for data insertion)
        if old_rows_count > 1:
            ws.delete_rows(start_row + 1, old_rows_count - 1)
    
    # === PHASE 3: Insert rows for new data ===
    rows_needed = len(invoices)
    
    if rows_needed > 1:
        ws.insert_rows(start_row + 1, rows_needed - 1)
    
    # === PHASE 4: Write data into cells (write BEFORE merging) ===
    from openpyxl.styles import Font
    for row_offset, inv in enumerate(invoices):
        current_row = start_row + row_offset
        
        # Write all individual cell values first
        ws.cell(row=current_row, column=1, value=row_offset + 1)
        ws.cell(row=current_row, column=2, value=inv['date'])
        ws.cell(row=current_row, column=3, value=inv['provider'])
        ws.cell(row=current_row, column=4, value=None)
        ws.cell(row=current_row, column=5, value=None)
        ws.cell(row=current_row, column=6, value=inv['amount'])
        ws.cell(row=current_row, column=6).number_format = '#,##0'
        ws.cell(row=current_row, column=7, value=inv['invoice_no'])
        
        # Apply styles from header row
        for col in range(1, 8):
            ref_cell = ws.cell(row=13, column=col)
            new_cell = ws.cell(row=current_row, column=col)
            if ref_cell.has_style:
                if ref_cell.font:
                    f = ref_cell.font
                    new_cell.font = Font(name=f.name, size=f.size, bold=False, italic=f.italic, color=f.color)
                if ref_cell.border:
                    new_cell.border = copy.copy(ref_cell.border)
                if ref_cell.alignment:
                    new_cell.alignment = copy.copy(ref_cell.alignment)
    
    # === PHASE 5: Apply merges AFTER all data is written ===
    for row_offset in range(rows_needed):
        current_row = start_row + row_offset
        ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=5)
                
    sum_row = start_row + rows_needed
    ws.cell(row=sum_row, column=6, value=f"=SUM(F{start_row}:F{sum_row-1})")
    
    for r in range(sum_row, sum_row + 10):
        cell = ws.cell(row=r, column=1)
        if cell.value and isinstance(cell.value, str) and "bằng chữ" in cell.value.lower():
            cell.value = f"(Viết bằng chữ: {number_to_words_vn(int(total_sum))})"
            break
            
    wb.save(output_path)
    print(f"Successfully saved to {output_path}")
    return output_path
